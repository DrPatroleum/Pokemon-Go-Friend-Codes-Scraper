import requests
from bs4 import BeautifulSoup
import re
import time, random
import qrcode
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

# =============================================================================
# SEKCJA: FUNKCJE SCRAPERA (pobieranie strony i parsowanie danych)
# =============================================================================
def fetch_page(url):
    """
    Pobiera zawartość strony pod danym adresem URL.
    W przypadku błędu wyświetla komunikat i zwraca None.
    """
    try:
        response = requests.get(url)
        response.raise_for_status()
        return response.text
    except requests.RequestException as e:
        print(f"Error fetching page: {e}")
        return None

def extract_friend_info(html):
    """
    Parsuje HTML, wyszukując informacje o znajomych z Pokémon GO.
    Zwraca listę słowników z danymi: name, level, code, location oraz team.
    """
    if not html:
        return []
    soup = BeautifulSoup(html, 'html.parser')
    friends = []
    # Szukamy elementów zawierających kod – bloki z klasą zawierającą "comment-bubble"
    for bubble in soup.find_all("div", class_=re.compile(r"comment-bubble")):
        # Nazwa trenera
        name_tag = bubble.find("span", class_="comment-bubble-header")
        name = name_tag.text.strip() if name_tag else "Unknown"
        
        # Poziom trenera
        level_match = re.search(r'Level\s*(\d+)', bubble.get_text())
        level = level_match.group(1) if level_match else "?"
        
        # Kod znajomego
        code_tag = bubble.find("strong")
        code = code_tag.text.strip() if code_tag else "NoCode"
        
        # Lokalizacja – wykluczamy linki prowadzące do profilu trenera
        comment_content = bubble.find("div", class_="comment-content")
        location_links = []
        if comment_content:
            for a in comment_content.find_all("a"):
                href = a.get("href", "")
                if not href.startswith("/trainer"):
                    location_links.append(a.text.strip())
        location = ", ".join(location_links) if location_links else "Unknown Location"
        
        # Pobranie informacji o teamie (Valor, Mystic, Instinct)
        team = "Unknown"
        if bubble.has_attr("class"):
            for cls in bubble.get("class"):
                cls_lower = cls.lower()
                if "valor" in cls_lower:
                    team = "Valor"
                    break
                elif "mystic" in cls_lower:
                    team = "Mystic"
                    break
                elif "instinct" in cls_lower:
                    team = "Instinct"
                    break
                    
        friend = {
            "name": name,
            "level": level,
            "code": code,
            "location": location,
            "team": team
        }
        friends.append(friend)
    return friends

# =============================================================================
# SEKCJA: FUNKCJE OBSŁUGI EXCELA (odczyt, zapis danych i formatowanie)
# =============================================================================
def read_existing_friend_codes(file_path):
    """
    Odczytuje istniejące kody z pliku Excel.
    Jeśli plik nie istnieje lub wystąpi błąd, zwraca pusty zbiór.
    """
    if not os.path.exists(file_path):
        return set()
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        friend_codes = set()
        # Pomijamy nagłówki (wiersz 1)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[2]:
                friend_codes.add(row[2])
        return friend_codes
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return set()

def write_new_friends(file_path, friends, qr_base_path):
    """
    Dodaje nowe rekordy do pliku Excel oraz wywołuje generowanie kodów QR.
    Nowe dane zawierają kolumny: Name, Level, Code, Location, Team oraz Date Added.
    Wiersze są formatowane kolorem w zależności od teamu.
    Operacje zapisu są zabezpieczone blokami try/except.
    """
    try:
        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            # Nagłówki kolumn
            sheet.append(["Name", "Level", "Code", "Location", "Team", "Date Added"])
    except Exception as e:
        print(f"Error opening or creating Excel file: {e}")
        return

    existing_codes = read_existing_friend_codes(file_path)
    
    # Definicje kolorów dla wypełnienia wierszy wg teamu
    team_colors = {
        "Valor": PatternFill(start_color="ff3b1f", end_color="ff3b1f", fill_type="solid"),     # czerwony
        "Mystic": PatternFill(start_color="00f1d7", end_color="00f1d7", fill_type="solid"),    # niebieski
        "Instinct": PatternFill(start_color="fde910", end_color="fde910", fill_type="solid")   # żółty
    }
    
    try:
        for friend in friends:
            if friend["code"] not in existing_codes:
                # Data dodania w formacie YYYY-MM-DD (bez godziny)
                date_added = datetime.now().strftime("%Y-%m-%d")
                row_data = [friend["name"], friend["level"], friend["code"], friend["location"], friend["team"], date_added]
                sheet.append(row_data)
                
                # Pobieramy numer ostatniego wiersza
                row_idx = sheet.max_row
                fill = team_colors.get(friend["team"], None)
                if fill:
                    for cell in sheet[row_idx]:
                        cell.fill = fill

                # Wyświetlamy informację o dodaniu bez teamu
                print(f"Added friend \033[91m{friend['name']}\033[0m || lvl \033[35m{friend['level']}\033[0m || \033[32m{friend['location']}\033[0m")
                # Generowanie kodu QR – identyfikator oparty o nazwę trenera
                generate_qr(friend["code"], friend["name"], qr_base_path)
                time.sleep(1)
        workbook.save(file_path)
    except Exception as e:
        print(f"Error writing to Excel file: {e}")

# =============================================================================
# SEKCJA: FUNKCJE GENEROWANIA QR (tworzenie unikalnej nazwy pliku i kodu QR)
# =============================================================================
def generate_unique_filename(trainer_name, qr_directory):
    """
    Generuje unikalną nazwę pliku na podstawie nazwy trenera.
    W przypadku konfliktu nazwy dodaje licznik.
    """
    base_name = trainer_name.replace(" ", "_")
    filename = f"{base_name}.jpg"
    counter = 1
    full_path = os.path.join(qr_directory, filename)
    while os.path.exists(full_path):
        filename = f"{base_name}_{counter}.jpg"
        full_path = os.path.join(qr_directory, filename)
        counter += 1
    return full_path

def generate_qr(data, trainer_name, base_qr_path):
    """
    Generuje kod QR dla przekazanych danych i zapisuje go w folderze "QR POGO CODES"
    w podanej ścieżce. Nazwa pliku oparta jest na nazwie trenera.
    """
    try:
        qr_directory = os.path.join(base_qr_path, "QR POGO CODES")
        if not os.path.exists(qr_directory):
            os.makedirs(qr_directory)
        qr_file = generate_unique_filename(trainer_name, qr_directory)
        
        qr = qrcode.QRCode(
            version=None,
            error_correction=qrcode.constants.ERROR_CORRECT_H,
            box_size=10,
            border=2
        )
        qr.add_data(data)
        qr.make(fit=True)
        img = qr.make_image(fill_color='black', back_color='white')
        img.save(qr_file)
    except Exception as e:
        print(f"Error generating QR for {trainer_name}: {e}")

# =============================================================================
# SEKCJA: KLASA POKEMONFRIENDSCRAPER (koordynacja całego procesu)
# =============================================================================
class PokemonFriendScraper:
    """
    Klasa zarządzająca procesem pobierania danych, parsowania,
    zapisu do Excela oraz generowania kodów QR.
    Parametry (URL, ścieżka do Excela, ścieżka podstawowa dla QR, zakres opóźnień)
    można konfigurować przy tworzeniu obiektu.
    """
    def __init__(self, url, excel_path, qr_base_path, refresh_interval_range=(120, 180)):
        self.url = url
        self.excel_path = excel_path
        self.qr_base_path = qr_base_path
        self.refresh_interval_range = refresh_interval_range

    def run(self):
        while True:
            # Wykonanie początkowego żądania (rozgrzewka)
            try:
                requests.get(self.url)
            except Exception as e:
                print(f"Initial GET request error: {e}")
            time.sleep(2)

            html = fetch_page(self.url)
            if html is None:
                print("Error fetching page content, skipping iteration.")
                time.sleep(random.uniform(*self.refresh_interval_range))
                continue

            friends = extract_friend_info(html)
            if friends:
                write_new_friends(self.excel_path, friends, self.qr_base_path)
            else:
                print("No friend info found.")

            total_friends = len(read_existing_friend_codes(self.excel_path))
            current_time = datetime.now().strftime("%H:%M:%S")
            print(f"\033[34m{current_time}\033[0m >>> \033[36m{total_friends}\033[0m Pokemon GO friends collected!")

            delay = random.uniform(*self.refresh_interval_range)
            time.sleep(delay)

# =============================================================================
# SEKCJA: URUCHOMIENIE APLIKACJI
# =============================================================================
if __name__ == "__main__":
    URL = "https://pokemongo.gishan.net/friends/codes/"
    FILE_PATH = "pokemon_friend_codes.xlsx"
    QR_PATH = r"C:\Users\USER\Desktop"
    
    scraper = PokemonFriendScraper(URL, FILE_PATH, QR_PATH)
    scraper.run()
